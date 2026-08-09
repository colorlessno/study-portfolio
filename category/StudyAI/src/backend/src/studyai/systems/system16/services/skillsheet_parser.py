from __future__ import annotations

import re
from io import BytesIO

from studyai.common.errors.models import ValidationAppError
from studyai.systems.system16.services.skill_normalizer import AliasRule, SkillNormalizer


class SkillsheetParser:
    HEADER_ROW_LIMIT = 6
    EXPERIENCE_PATTERN = re.compile(r"(\d+(?:\.\d+)?)\s*年")
    PROJECT_PATTERN = re.compile(r"(案件|プロジェクト|project)")

    def __init__(self) -> None:
        self.normalizer = SkillNormalizer()

    def parse_skillsheet(self, file_name: str, file_bytes: bytes, rules: list[AliasRule]) -> dict:
        if not file_name.lower().endswith(".xlsx"):
            raise ValidationAppError("invalid_skillsheet_format", "The skillsheet must be an .xlsx file.")
        rows = self._extract_rows(file_bytes)
        if not rows:
            raise ValidationAppError("empty_skillsheet", "The skillsheet is empty.")

        layout_type = self._detect_layout(rows)
        text_blob = "\n".join(" | ".join(row) for row in rows if row)
        catalog = self.normalizer.extract_catalog(text_blob, rules)

        total_projects = self._estimate_total_projects(rows)
        total_years = self._estimate_total_years(text_blob)
        unresolved_skills = self._find_unresolved_skill_tokens(text_blob, catalog["technical_all"])

        review_reasons: list[str] = []
        if layout_type == "review_required":
            review_reasons.append("unsupported_skillsheet_layout")
        if total_projects == 0:
            review_reasons.append("project_history_not_detected")
        if len(catalog["technical_all"]) < 3:
            review_reasons.append("too_few_detected_skills")
        if unresolved_skills:
            review_reasons.append("unresolved_skills_present")

        parse_confidence = self._calculate_parse_confidence(
            layout_type=layout_type,
            total_projects=total_projects,
            detected_skill_count=len(catalog["technical_all"]),
            unresolved_count=len(unresolved_skills),
        )
        review_required = parse_confidence < 0.75 or bool(unresolved_skills)

        return {
            "layout_type": layout_type,
            "parse_confidence": parse_confidence,
            "review_required": review_required,
            "review_reasons": review_reasons,
            "unresolved_skills": unresolved_skills,
            "parsed_result": {
                "total_projects": total_projects,
                "total_experience_years": total_years,
                "skills": {
                    "languages": catalog["languages"],
                    "databases": catalog["databases"],
                    "os": catalog["os"],
                    "tools": catalog["tools"],
                },
                "processes": catalog["processes"],
                "roles": catalog["roles"],
                "domains": catalog["domains"],
            },
        }

    def _extract_rows(self, file_bytes: bytes) -> list[list[str]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValidationAppError("xlsx_support_missing", "openpyxl is not installed.") from exc

        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        rows: list[list[str]] = []
        try:
            for sheet in workbook.worksheets:
                rows.append([f"[sheet] {sheet.title}"])
                for row in sheet.iter_rows(values_only=True):
                    normalized = [str(value).strip() for value in row if value is not None and str(value).strip()]
                    if normalized:
                        rows.append(normalized)
        finally:
            workbook.close()
        return rows

    def _detect_layout(self, rows: list[list[str]]) -> str:
        header_text = "\n".join(" | ".join(row).casefold() for row in rows[: self.HEADER_ROW_LIMIT])
        if "スキル" in header_text or "経験年数" in header_text or "technical" in header_text:
            return "A"
        if "案件" in header_text or "プロジェクト" in header_text or "担当" in header_text:
            return "B"
        return "review_required"

    def _estimate_total_projects(self, rows: list[list[str]]) -> int:
        count = 0
        for row in rows:
            joined = " ".join(row).casefold()
            if self.PROJECT_PATTERN.search(joined):
                count += 1
        return max(count, 1 if rows else 0)

    def _estimate_total_years(self, text_blob: str) -> float:
        matches = [float(match.group(1)) for match in self.EXPERIENCE_PATTERN.finditer(text_blob)]
        if not matches:
            return 0.0
        return round(max(matches), 1)

    def _find_unresolved_skill_tokens(self, text_blob: str, recognized_skills: list[str]) -> list[str]:
        candidates = re.findall(r"[A-Za-z][A-Za-z0-9#+.-]{2,}", text_blob)
        blocked = {token.casefold() for token in recognized_skills}
        blocked.update({"sheet", "project", "skill", "experience", "server", "client"})
        unresolved = []
        for token in candidates:
            lowered = token.casefold()
            if lowered in blocked:
                continue
            if lowered not in unresolved:
                unresolved.append(lowered)
        return unresolved[:5]

    @staticmethod
    def _calculate_parse_confidence(
        *,
        layout_type: str,
        total_projects: int,
        detected_skill_count: int,
        unresolved_count: int,
    ) -> float:
        score = 0.45
        if layout_type in {"A", "B"}:
            score += 0.20
        if total_projects > 0:
            score += 0.10
        score += min(detected_skill_count, 6) * 0.04
        score -= min(unresolved_count, 4) * 0.04
        return round(max(0.0, min(1.0, score)), 3)
